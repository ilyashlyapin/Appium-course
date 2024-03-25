import time

import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.appium_service import AppiumService
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver import ActionChains

from testcases.scroll_util import ScrollUtil
from utilities import dataProvider


@pytest.fixture
def log_on_failure(request,appium_driver):
    yield
    item = request.node
    if item.rep_call.failed:
        allure.attach(driver.get_screenshot_as_png(), name="screenshot", attachment_type=AttachmentType.PNG)

def get_data():
    # return [
    #     ("Мама"),
    #     ("Отеc")
    # ]
    workbook = openpyxl.load_workbook("..//excel/testdata.xlsx")
    sheet = workbook["SearchTest"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []
    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1,totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
    return mainList

# no array inside array, only tuple inside array


def setup_function():
    global appium_service
    appium_service = AppiumService()
    options = UiAutomator2Options()
    options.platformVersion = '13'
    options.udid = 'b4c0762b'
    options.app_package = 'com.google.android.contacts'
    options.app_activity = 'com.google.android.apps.contacts.activities.PeopleActivity'
    options.auto_grant_permissions = True
    appium_service.start()
    global driver
    driver = webdriver.Remote('http://127.0.0.1:4723', options=options)
    driver.implicitly_wait(5)


def teardown_function():
    driver.quit()
    appium_service.stop()


@pytest.mark.parametrize("name", get_data())
def test_longpressname(name, log_on_failure):
    # print("the name is " + name)
    name_text = driver.find_element(AppiumBy.ACCESSIBILITY_ID, name).text

#     print(name_text)x
    assert name_text == name
    # allure.attach(driver.get_screenshot_as_png(),name="screenshot",attachment_type=AttachmentType.PNG)
# #
# @pytest.skip
# @pytest.mark.usefixtures("log_on_failure")
# @pytest.mark.parametrize("city,country", dataProvider.get_data("SearchTest"))
# def test_dologin(city,country,appium_driver):
#     driver = appium_driver
#     driver.find_element_by_id('com.goibibo:id/btn1').click()
#     driver.find_element_by_accessibility_id('destination').click()
#     driver.find_element_by_id('com.goibibo:id/edtSearch').send_keys(city)
#     driver.find_elements_by_id('com.goibibo:id/lytLocationItem')[0].click()
#     driver.find_element_by_xpath("//android.widget.TextView[@text='Search']").click()
#     cityText = driver.find_element_by_xpath("//android.widget.TextView[contains(@text,'EXPLORE')]").text
#     print(cityText)
#     newCityText = str(cityText).replace("EXPLORE ","").replace("!","")
#     print(newCityText)
#
#     assert newCityText in str(city).upper()
#     #allure.attach(driver.get_screenshot_as_png(),name="screenshot",attachment_type=AttachmentType.PNG)


#
# def test_searchVillas(appium_driver):
#     driver = appium_driver
#     driver.find_element_by_xpath("//android.widget.TextView[@text='Villas & Apts']").click()
#     driver.find_element_by_xpath("//android.widget.TextView[@text='Area, Landmark or Property']").click()
#     driver.find_element_by_id('com.goibibo:id/edtSearch').send_keys("Delhi")
#     driver.hide_keyboard()
# driver.find_elements_by_id('com.goibibo:id/lytLocationItem')[0].click()
# driver.find_element_by_xpath("//android.widget.TextView[@text='6']").click()
# driver.find_element_by_xpath("//android.widget.TextView[@text='11']").click()
# driver.find_element_by_xpath("//android.widget.TextView[contains(@text,'Continue')]").click()
# driver.find_element_by_xpath("//android.widget.TextView[contains(@text,'View Stays')]").click()
# ScrollUtil.scrollToTextByAndroidUIAutomator("Shubham Vilas", driver)
